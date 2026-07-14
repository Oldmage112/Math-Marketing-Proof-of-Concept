#!/usr/bin/env python3
"""
download_commons_assets.py

Reads the "Wikimedia Commons URL" column out of one or more catalog spreadsheets
(the .xlsx files Claude generated) and downloads the actual media files into a
local folder, organized into subfolders by category.

Works for both the image catalog and the animation/video catalog, since they
share the same column layout (Category, File Name, Description,
Wikimedia Commons URL, License, Source).

-------------------------------------------------------------------------------
SETUP (run once)
-------------------------------------------------------------------------------
    

-------------------------------------------------------------------------------
USAGE
-------------------------------------------------------------------------------
    # Download everything from one catalog into ./downloads
    python download_commons_assets.py Public_Domain_Math_Assets_Catalog.xlsx

    # Download from both catalogs at once, into a custom folder
    python download_commons_assets.py Public_Domain_Math_Assets_Catalog.xlsx \
        Public_Domain_Math_Animations_Catalog.xlsx -o my_math_media

    # Preview what would be downloaded without actually downloading anything
    python download_commons_assets.py Public_Domain_Math_Assets_Catalog.xlsx --dry-run

    # Skip anything bigger than 50 MB (handy for the video catalog)
    python download_commons_assets.py Public_Domain_Math_Animations_Catalog.xlsx --max-size-mb 50

Re-running the script skips files that were already downloaded, so it's safe
to stop and resume.
-------------------------------------------------------------------------------
"""

import argparse
import csv
import re
import sys
import time
import urllib.parse
from pathlib import Path
from typing import Iterator, Optional, Tuple

try:
    import requests
except ImportError:
    sys.exit("Missing dependency 'requests'. Install it with:  pip install requests")

try:
    from openpyxl import load_workbook
except ImportError:
    sys.exit("Missing dependency 'openpyxl'. Install it with:  pip install openpyxl")


# Wikimedia asks bots/scripts to identify themselves with a descriptive
# User-Agent. Feel free to edit the contact info below.
USER_AGENT = (
    "MathAssetDownloaderScript/1.0 "
    "(personal/educational use; contact: replace-with-your-email@example.com) "
    "python-requests"
)

REQUEST_TIMEOUT = 30      # seconds
RETRY_COUNT = 3
RETRY_DELAY = 3           # seconds between retries
DEFAULT_THROTTLE = 1.0    # seconds between downloads, to be polite to Commons


def extract_filename_from_commons_url(url: str) -> str:
    """
    Given a Wikimedia Commons file page URL like
    'https://commons.wikimedia.org/wiki/File:Fractal_Broccoli.jpg',
    return the decoded file name 'Fractal Broccoli.jpg'.
    """
    marker = "/wiki/File:"
    idx = url.find(marker)
    if idx == -1:
        raise ValueError(f"Not a recognizable Commons File: URL: {url}")
    raw_name = url[idx + len(marker):]
    raw_name = urllib.parse.unquote(raw_name)
    raw_name = raw_name.replace("_", " ")
    return raw_name


def special_filepath_url(filename: str) -> str:
    """
    Build the Special:FilePath URL for a Commons file name. Requesting this
    URL 302-redirects to the actual file on upload.wikimedia.org, so it works
    as a generic "give me the current version of this file" endpoint.
    """
    return "https://commons.wikimedia.org/wiki/Special:FilePath/" + urllib.parse.quote(filename)


def safe_filename(name: str) -> str:
    """Strip characters that are illegal in Windows/Mac/Linux file names."""
    cleaned = re.sub(r'[<>:"/\\|?*]', "_", name).strip()
    return cleaned or "unnamed_file"


def find_column(header_row, *keywords) -> Optional[int]:
    """Return the index of the first header cell containing all keywords (case-insensitive)."""
    for i, cell in enumerate(header_row):
        text = str(cell or "").strip().lower()
        if all(k in text for k in keywords):
            return i
    return None


def read_catalog(xlsx_path: Path) -> Iterator[Tuple[str, str, str]]:
    """Yield (category, file_name, commons_url) tuples from a catalog spreadsheet."""
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        return

    header = rows[0]
    cat_idx = find_column(header, "categ")
    name_idx = find_column(header, "file", "name")
    url_idx = find_column(header, "url")

    if url_idx is None:
        print(f"  ! Could not find a URL column in {xlsx_path.name} - skipping this file.")
        return

    for row in rows[1:]:
        if not row or all(v is None for v in row):
            continue
        category = str(row[cat_idx]).strip() if cat_idx is not None and row[cat_idx] else "Uncategorized"
        file_name = str(row[name_idx]).strip() if name_idx is not None and row[name_idx] else ""
        url = str(row[url_idx]).strip() if row[url_idx] else ""
        if url:
            yield category, file_name, url


def remote_size_mb(session: requests.Session, file_url: str) -> Optional[float]:
    """Best-effort HEAD request to check file size before downloading. Returns None if unknown."""
    try:
        resp = session.head(file_url, timeout=REQUEST_TIMEOUT, allow_redirects=True)
        length = resp.headers.get("Content-Length")
        if length:
            return int(length) / (1024 * 1024)
    except requests.RequestException:
        pass
    return None


def download_file(
    session: requests.Session,
    commons_url: str,
    dest_path: Path,
    max_size_mb: Optional[float],
    dry_run: bool,
) -> Tuple[str, str]:
    """
    Download one file via Special:FilePath, with retries.
    Returns (status, message) where status is one of:
    'ok', 'skipped', 'too_large', 'error'
    """
    if dest_path.exists() and dest_path.stat().st_size > 0:
        return "skipped", "already exists on disk"

    try:
        filename = extract_filename_from_commons_url(commons_url)
    except ValueError as e:
        return "error", str(e)

    file_url = special_filepath_url(filename)

    if max_size_mb is not None:
        size = remote_size_mb(session, file_url)
        if size is not None and size > max_size_mb:
            return "too_large", f"{size:.1f} MB > limit of {max_size_mb} MB"

    if dry_run:
        return "ok", "would download (dry run)"

    last_error = ""
    for attempt in range(1, RETRY_COUNT + 1):
        try:
            resp = session.get(file_url, timeout=REQUEST_TIMEOUT, stream=True, allow_redirects=True)
            if resp.status_code == 200:
                dest_path.parent.mkdir(parents=True, exist_ok=True)
                tmp_path = dest_path.with_suffix(dest_path.suffix + ".part")
                with open(tmp_path, "wb") as f:
                    for chunk in resp.iter_content(chunk_size=1024 * 256):
                        if chunk:
                            f.write(chunk)
                tmp_path.rename(dest_path)
                return "ok", "downloaded"
            elif resp.status_code == 404:
                return "error", f"404 not found ({file_url})"
            else:
                last_error = f"HTTP {resp.status_code}"
        except requests.RequestException as e:
            last_error = str(e)
        if attempt < RETRY_COUNT:
            time.sleep(RETRY_DELAY)

    return "error", f"failed after {RETRY_COUNT} attempts: {last_error}"


def main():
    parser = argparse.ArgumentParser(
        description="Download Wikimedia Commons files listed in Claude-generated catalog spreadsheets.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument("catalogs", nargs="+", help="One or more .xlsx catalog files")
    parser.add_argument("-o", "--output", default="downloads", help="Output folder (default: ./downloads)")
    parser.add_argument("--throttle", type=float, default=DEFAULT_THROTTLE,
                         help=f"Seconds to wait between downloads (default: {DEFAULT_THROTTLE})")
    parser.add_argument("--max-size-mb", type=float, default=None,
                         help="Skip files larger than this size in MB (checked via HEAD request)")
    parser.add_argument("--dry-run", action="store_true",
                         help="Preview what would be downloaded without saving any files")
    args = parser.parse_args()

    out_root = Path(args.output)
    out_root.mkdir(parents=True, exist_ok=True)

    session = requests.Session()
    session.headers.update({"User-Agent": USER_AGENT})

    log_rows = []
    counts = {"ok": 0, "skipped": 0, "too_large": 0, "error": 0}
    total = 0

    for catalog_arg in args.catalogs:
        catalog_path = Path(catalog_arg)
        if not catalog_path.exists():
            print(f"! Skipping missing file: {catalog_path}")
            continue

        print(f"\n=== Reading {catalog_path.name} ===")
        entries = list(read_catalog(catalog_path))
        print(f"  Found {len(entries)} entries")

        for category, file_name, commons_url in entries:
            total += 1
            cat_folder = safe_filename(category)
            base_name = safe_filename(file_name) if file_name else safe_filename(
                extract_filename_from_commons_url(commons_url)
            )
            dest = out_root / cat_folder / base_name

            print(f"  [{total}] {cat_folder}/{base_name} ...", end=" ", flush=True)
            try:
                status, msg = download_file(session, commons_url, dest, args.max_size_mb, args.dry_run)
            except KeyboardInterrupt:
                print("\nInterrupted by user.")
                raise
            print(f"{status.upper()} - {msg}")

            counts[status] = counts.get(status, 0) + 1
            log_rows.append({
                "catalog": catalog_path.name,
                "category": category,
                "file_name": base_name,
                "commons_url": commons_url,
                "status": status,
                "detail": msg,
            })

            if status == "ok" and not args.dry_run:
                time.sleep(args.throttle)

    log_path = out_root / "_download_log.csv"
    with open(log_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(
            f, fieldnames=["catalog", "category", "file_name", "commons_url", "status", "detail"]
        )
        writer.writeheader()
        writer.writerows(log_rows)

    print("\n" + "=" * 50)
    print(f"Done. {total} entries processed.")
    print(f"  downloaded : {counts.get('ok', 0)}")
    print(f"  skipped    : {counts.get('skipped', 0)} (already on disk)")
    print(f"  too large  : {counts.get('too_large', 0)}")
    print(f"  errors     : {counts.get('error', 0)}")
    print(f"Log written to: {log_path}")
    if counts.get("error", 0):
        print("\nCheck the log for error details - a common cause is a file having")
        print("been renamed or deleted on Commons since the catalog was built.")


if __name__ == "__main__":
    main()
